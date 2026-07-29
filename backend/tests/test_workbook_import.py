from io import BytesIO
from unittest.mock import Mock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook as ExcelWorkbook
from sqlalchemy import func, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.main import app
from app.models import Flashcard, StudySheet, Workbook
from app.services.excel_types import ParsedCard, ParsedSheet, ParsedWorkbook
from app.services.workbook_import import (
    WorkbookImportPersistenceError,
    import_parsed_workbook,
)


def workbook_bytes(*sheets: tuple[str, list[tuple[object, ...]]]) -> BytesIO:
    workbook = ExcelWorkbook()
    first_sheet = workbook.active

    for index, (title, rows) in enumerate(sheets):
        worksheet = first_sheet if index == 0 else workbook.create_sheet()
        worksheet.title = title
        for row in rows:
            worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    return buffer


@pytest.fixture()
def api_client(db_session: Session):
    def override_get_db():
        yield db_session

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as client:
        yield client
    app.dependency_overrides.clear()


def test_import_service_persists_relationships_and_aggregates(db_session: Session) -> None:
    parsed = ParsedWorkbook(
        sheets=[
            ParsedSheet(
                sheet_name="Part 1",
                position=1,
                cards=[
                    ParsedCard(1, "follow up", "theo dõi", "Follow up tomorrow.", None),
                    ParsedCard(2, "look into", "xem xét", None, "Tôi sẽ xem xét."),
                ],
            ),
            ParsedSheet(
                sheet_name="Part 2",
                position=2,
                cards=[ParsedCard(1, "prepare", "chuẩn bị", None, None)],
            ),
        ]
    )

    workbook = import_parsed_workbook(db_session, parsed, "Interview Vocabulary.xlsx")
    db_session.expire_all()
    stored_workbook = db_session.get(Workbook, workbook.id)

    assert stored_workbook is not None
    assert stored_workbook.name == "Interview Vocabulary"
    assert stored_workbook.sheet_count == 2
    assert stored_workbook.total_cards == 3
    assert [sheet.position for sheet in stored_workbook.sheets] == [1, 2]
    assert [sheet.card_count for sheet in stored_workbook.sheets] == [2, 1]
    assert stored_workbook.sheets[0].flashcards[0].example_vi is None
    assert stored_workbook.sheets[0].flashcards[1].example_en is None


def test_import_service_rolls_back_after_database_error() -> None:
    db = Mock()
    db.commit.side_effect = SQLAlchemyError("database unavailable")
    parsed = ParsedWorkbook(sheets=[])

    with pytest.raises(WorkbookImportPersistenceError):
        import_parsed_workbook(db, parsed, "Vocabulary.xlsx")

    db.add.assert_called_once()
    db.rollback.assert_called_once()


def test_import_endpoint_creates_workbook_from_xlsx(api_client: TestClient, db_session: Session) -> None:
    source = workbook_bytes(
        (
            "Part 1",
            [
                ("Phrase", "Meaning", "Example EN", "Example VI"),
                ("follow up", "theo dõi", "Follow up tomorrow.", None),
                ("look into", "xem xét", None, "Tôi sẽ xem xét."),
            ],
        ),
        (
            "Part 2",
            [
                ("Phrase", "Meaning"),
                ("prepare", "chuẩn bị"),
            ],
        ),
    )

    response = api_client.post(
        "/api/v1/workbooks/import",
        files={"file": ("Interview.XLSX", source, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    assert response.json() == {
        "id": 1,
        "name": "Interview",
        "original_filename": "Interview.XLSX",
        "sheet_count": 2,
        "total_cards": 3,
        "imported_at": response.json()["imported_at"],
        "sheets": [
            {"id": 1, "name": "Part 1", "position": 1, "card_count": 2},
            {"id": 2, "name": "Part 2", "position": 2, "card_count": 1},
        ],
    }
    assert db_session.scalar(select(func.count(Workbook.id))) == 1
    assert db_session.scalar(select(func.count(StudySheet.id))) == 2
    assert db_session.scalar(select(func.count(Flashcard.id))) == 3


def test_import_endpoint_rejects_wrong_extension_without_database_write(
    api_client: TestClient, db_session: Session
) -> None:
    response = api_client.post(
        "/api/v1/workbooks/import",
        files={"file": ("Vocabulary.csv", BytesIO(b"phrase,meaning"), "text/csv")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Only .xlsx workbook files are supported."
    assert db_session.scalar(select(func.count(Workbook.id))) == 0


def test_import_endpoint_rejects_overlong_workbook_name(api_client: TestClient) -> None:
    response = api_client.post(
        "/api/v1/workbooks/import",
        files={"file": (f"{'a' * 256}.xlsx", BytesIO(b"unused"), "application/octet-stream")},
    )

    assert response.status_code == 422
    assert response.json()["detail"] == "Workbook filename must be at most 255 characters."


def test_import_endpoint_returns_validation_errors_without_partial_write(
    api_client: TestClient, db_session: Session
) -> None:
    source = workbook_bytes(
        (
            "Vocabulary",
            [
                ("Phrase", "Meaning"),
                ("valid", "hợp lệ"),
                ("missing meaning", None),
            ],
        ),
    )

    response = api_client.post(
        "/api/v1/workbooks/import",
        files={"file": ("Invalid.xlsx", source, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 422
    assert response.json()["detail"] == [
        {
            "sheet_name": "Vocabulary",
            "row_number": 3,
            "column": "Meaning",
            "message": "Required value is missing.",
        }
    ]
    assert db_session.scalar(select(func.count(Workbook.id))) == 0


def test_import_endpoint_rejects_unreadable_xlsx_without_database_write(
    api_client: TestClient, db_session: Session
) -> None:
    response = api_client.post(
        "/api/v1/workbooks/import",
        files={"file": ("Broken.xlsx", BytesIO(b"not a workbook"), "application/octet-stream")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "The uploaded file is not a readable .xlsx workbook."
    assert db_session.scalar(select(func.count(Workbook.id))) == 0


def test_import_endpoint_allows_post_cors_preflight(api_client: TestClient) -> None:
    response = api_client.options(
        "/api/v1/workbooks/import",
        headers={
            "Origin": "http://localhost:3000",
            "Access-Control-Request-Method": "POST",
            "Access-Control-Request-Headers": "content-type",
        },
    )

    assert response.status_code == 200
    assert response.headers["access-control-allow-origin"] == "http://localhost:3000"
    assert "POST" in response.headers["access-control-allow-methods"]
    assert "content-type" in response.headers["access-control-allow-headers"].lower()
