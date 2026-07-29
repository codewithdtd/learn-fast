from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel_types import (
    ExcelParseError,
    ExcelValidationError,
    ParsedCard,
    ParsedSheet,
    ParsedWorkbook,
    ValidationErrorDetail,
)


HEADER_KEYS = {
    "phrase": "phrase",
    "meaning": "meaning",
    "example en": "example_en",
    "example vi": "example_vi",
}
REQUIRED_HEADERS = {"phrase": "Phrase", "meaning": "Meaning"}


def normalize_header(value: object) -> str:
    """Normalize Excel header casing and whitespace before matching known columns."""
    return " ".join(str(value or "").strip().lower().split())


def normalize_cell_text(value: object) -> str | None:
    """Return trimmed text while preserving non-empty numeric and boolean cell values."""
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def parse_excel_workbook(source: BinaryIO) -> ParsedWorkbook:
    """Parse and validate an .xlsx workbook without writing to the database."""
    try:
        source.seek(0)
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as error:
        raise ExcelParseError("The uploaded file is not a readable .xlsx workbook.") from error

    errors: list[ValidationErrorDetail] = []
    parsed_sheets: list[ParsedSheet] = []

    try:
        for sheet_position, worksheet in enumerate(workbook.worksheets, start=1):
            if _is_sheet_empty(worksheet):
                continue

            header_mapping = _build_header_mapping(worksheet, errors)
            if header_mapping is None:
                continue

            cards = _parse_cards(worksheet, header_mapping, errors)
            parsed_sheets.append(
                ParsedSheet(
                    sheet_name=worksheet.title,
                    position=sheet_position,
                    cards=cards,
                )
            )
    finally:
        workbook.close()

    if errors:
        # Returning all issues lets the import UI show every fix needed instead
        # of forcing the user to upload the same workbook repeatedly.
        raise ExcelValidationError(errors)

    return ParsedWorkbook(sheets=parsed_sheets)


def _is_sheet_empty(worksheet: Worksheet) -> bool:
    return not any(
        normalize_cell_text(value) is not None
        for row in worksheet.iter_rows(values_only=True)
        for value in row
    )


def _build_header_mapping(
    worksheet: Worksheet,
    errors: list[ValidationErrorDetail],
) -> dict[str, int] | None:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    mapping: dict[str, int] = {}

    for index, value in enumerate(header_row):
        known_key = HEADER_KEYS.get(normalize_header(value))
        if known_key is None:
            continue

        if known_key in mapping:
            errors.append(
                ValidationErrorDetail(
                    sheet_name=worksheet.title,
                    row_number=1,
                    column=_display_column(known_key),
                    message="Duplicate recognized header after normalization.",
                )
            )
            continue

        mapping[known_key] = index

    for key, display_name in REQUIRED_HEADERS.items():
        if key not in mapping:
            errors.append(
                ValidationErrorDetail(
                    sheet_name=worksheet.title,
                    row_number=1,
                    column=display_name,
                    message="Required header is missing.",
                )
            )

    return mapping if not _has_header_errors(worksheet.title, errors) else None


def _parse_cards(
    worksheet: Worksheet,
    header_mapping: dict[str, int],
    errors: list[ValidationErrorDetail],
) -> list[ParsedCard]:
    cards: list[ParsedCard] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {
            key: normalize_cell_text(row[index]) if index < len(row) else None
            for key, index in header_mapping.items()
        }
        if not any(values.values()):
            continue

        missing_required = [key for key in REQUIRED_HEADERS if values.get(key) is None]
        for key in missing_required:
            errors.append(
                ValidationErrorDetail(
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    column=_display_column(key),
                    message="Required value is missing.",
                )
            )

        if missing_required:
            continue

        cards.append(
            ParsedCard(
                position=len(cards) + 1,
                phrase=values["phrase"],
                meaning=values["meaning"],
                example_en=values.get("example_en"),
                example_vi=values.get("example_vi"),
            )
        )

    return cards


def _has_header_errors(sheet_name: str, errors: list[ValidationErrorDetail]) -> bool:
    return any(error.sheet_name == sheet_name and error.row_number == 1 for error in errors)


def _display_column(key: str) -> str:
    return {
        "phrase": "Phrase",
        "meaning": "Meaning",
        "example_en": "Example EN",
        "example_vi": "Example VI",
    }[key]
